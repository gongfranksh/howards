# -*- coding: utf-8 -*-
import decimal
import hashlib
import json
import os
import sys
import openpyxl

from export_files import export_file
from howarddb import ScopedSession
from howarddb.bean_baseinfo import BaseInfo, query_same_baseinfo, match_baseinfo, ApplyLine
from howarddb.howard_db import columns_to_dict
from make_tags import get_keyword
from utils import RMB2USD, SOURCE_BLANK_LINES

reload(sys)
# sys.setdefaultencoding('utf8')
sys.setdefaultencoding('gbk')


def scan_work_folder():
    for root, dirs, files in os.walk("./in"):
        # print('root_dir:', root)  # 当前目录路径
        # print('sub_dirs:', dirs)  # 当前路径下所有子目录
        # print('files:', files)  # 当前路径下所有非目录子文件

        for file in files:
            print os.path.join(root, file)
            make_apply_files(file,root)


def make_apply_files(filename,dirs):
        # print(dirs)
        filename_no_extend=filename.replace(".xlsx","")
        wb = openpyxl.load_workbook(dirs+'/'+filename,data_only=True)
        for ws in wb:
            rows = []
            need_apply=[]
            tmp_total_weight=decimal.Decimal(0.00)
            i = 0
            for row in ws.iter_rows():
               if i<>0:
                   rows.append(row)
               i+=1
            for row in rows:
                item_name=''
                item_qty=0
                item_brand=None
                rules=None

                for col in row:
                        if col.column==1:
                            # print(col.value)
                            item_name=col.value
                            keyword_list=get_keyword(col.value)
                            rules=match_baseinfo(keyword_list)
                            if rules is not None:
                                print rules.id

                        if col.column==2:
                            item_qty=col.value
                            print(col.value)

                        if col.column==3:
                            item_brand=col.value
                            print(col.value)

                line=ApplyLine()
                line.brand=item_brand
                line.name=item_name
                line.qty=item_qty
                line.unit=rules.unit
                line.weight=rules.weight
                line.hs=rules.tax_code
                line.amount=rules.amount
                line.name_english=rules.name_english
                line.english_elements=rules.english_elements
                line.unit_english=rules.unit_english
                line.weight_total_net=rules.weight * item_qty
                line.amount_subtotal=rules.amount * item_qty
                line.amount_subtotal_tax_rmb=rules.general_rate * rules.amount * item_qty * RMB2USD
                tmp_total_weight=tmp_total_weight+line.weight_total_net

                need_apply.append(line)

            export={
                'apply':need_apply,
                'filename':filename_no_extend+'_'+ws.title,
                'input_net_weight': ws.cell(row=1, column=6).value,
                'sum_total_weight': tmp_total_weight
            }

            #判断excel工作表不为空才能导出
            if  (ws.max_column+ ws.max_row) > SOURCE_BLANK_LINES:
                export_file(export)

# make_apply_files()
scan_work_folder()