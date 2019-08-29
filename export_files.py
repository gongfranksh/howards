# -*- coding: utf-8 -*-
import datetime
import decimal
import hashlib
import json
import sys
from datetime import time
from shutil import copyfile

import openpyxl
from openpyxl import load_workbook

from howarddb import ScopedSession
from howarddb.bean_baseinfo import BaseInfo, query_same_baseinfo
from howarddb.howard_db import columns_to_dict
from utils import OUT_PATH_PRFIX, OUT_FILE_EXT, OUT_TEMPLATE_HEAD_LINES, OUT_TEMPLATE_FILE_NAME

reload(sys)
sys.setdefaultencoding('utf8')

def export_file(inexport):
    try:
        if inexport is not None:
            itemlines=inexport['apply']
            if inexport['sum_total_weight'] is not None:
                item_total_weight=inexport['sum_total_weight']
            else:
                item_total_weight=decimal.Decimal(0.00)

            if inexport['input_net_weight'] is not None:
                item_input_weight=inexport['input_net_weight']
            else:
                item_input_weight=decimal.Decimal(0.00)

            itemname=OUT_PATH_PRFIX+inexport['filename']+'_'\
                     +datetime.datetime.now().strftime('%Y%m%d%H%M%S')\
                     +OUT_FILE_EXT
            copy_from_template(itemname)
            insert_excel(itemname,itemlines,item_total_weight,item_input_weight)
    except Exception, e:
        print(e.message)
        return None

def copy_from_template(filename):
    copyfile(OUT_TEMPLATE_FILE_NAME,filename)

def insert_excel(filename,items,total_weight,input_weight):
    wb = load_workbook(filename)
    ws=wb[u'申报']
    i=OUT_TEMPLATE_HEAD_LINES
    for item in items:
        ws.cell(row=i, column=1).value = i-2
        ws.cell(row=i, column=2).value = item.hs
        ws.cell(row=i, column=3).value = item.name_english
        ws.cell(row=i, column=5).value = item.qty
        ws.cell(row=i, column=6).value = item.unit_english
        ws.cell(row=i, column=7).value = item.amount
        ws.cell(row=i, column=8).value = item.amount_subtotal
        ws.cell(row=i, column=9).value = item.weight_total_net

        if total_weight <> 0:
            ocr_rate=item.weight_total_net / total_weight

        ws.cell(row=i, column=10).value = ocr_rate * input_weight
        ws.cell(row=i, column=11).value = item.brand
        ws.cell(row=i, column=13).value = item.english_elements
        i+=1
    wb.save(filename)

