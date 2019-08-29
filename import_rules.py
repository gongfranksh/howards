# -*- coding: utf-8 -*-
import hashlib
import json
import sys
import openpyxl

from howarddb import ScopedSession
from howarddb.bean_baseinfo import BaseInfo, query_same_baseinfo
from howarddb.howard_db import columns_to_dict

reload(sys)
sys.setdefaultencoding('utf8')



def input_rules():
        wb = openpyxl.load_workbook('./excel/datatrans.xlsx',data_only=True)
        # 从工作薄中获取一个表单(sheet)对象
        # sheets = wb.sheetnames
        ws = wb[u'基础数据']

        rows = []
        i=0

        se=ScopedSession()
        for row in ws.iter_rows():
           if i<>0:
               rows.append(row)
           i+=1

        for row in rows:
            binfo = BaseInfo()
            for col in row:
                    if col.column==1:
                        binfo.name=col.value
                    if col.column==2:
                        binfo.tax_code=col.value
                    if col.column==3:
                        binfo.cust_rate=col.value
                    if col.column==4:
                        binfo.vat_rate=col.value
                    if col.column==5:
                        binfo.general_rate=col.value
                    if col.column==6:
                        binfo.amount=col.value
                    if col.column==7:
                        binfo.weight=col.value
                    if col.column==8:
                        binfo.unit=col.value
                    if col.column==9:
                        binfo.english_elements=col.value
                    if col.column==10:
                        binfo.apply_elements=col.value
                    if col.column==11:
                        binfo.name_english=col.value
                    if col.column==12:
                        binfo.unit_english=col.value

            #去掉重复的规则
            str_json=json.dumps(columns_to_dict(binfo))
            str_md5=hashlib.md5(str_json.encode(encoding='UTF-8')).hexdigest()
            binfo.md5=str_md5
            if query_same_baseinfo(se,str_md5):
                se.add(binfo)

        se.flush()
        se.commit()

input_rules()